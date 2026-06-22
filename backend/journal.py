"""
Génération du journal d'écritures comptables (format Cegid Quadra).

À partir des relevés nettoyés (TikTok/Shopify), produit pour **chaque jour** une
écriture en **partie double équilibrée** (Débit = Crédit), prête à importer dans
Quadra. Calqué sur le modèle fourni `MODELE JNL TIKTOK FRANCE.xlsx`.

Principe (vérifié sur le modèle, au centime) :
  - Débit  90xxxxxx  = total_settlement (l'encaissement net reçu de la plateforme)
  - Débit  622…      = commission HT (frais plateforme)
  - Débit  44566…    = TVA déductible sur la commission
  - Crédit 707…      = ventes HT
  - Crédit 70800…    = port HT
  - Crédit 44572…    = TVA collectée sur (ventes + port)

L'écriture s'équilibre par construction car, dans les données,
`total_settlement ≈ net_sales + shipping + fees` (fees négatifs). Tout écart
résiduel d'arrondi (≤ 0,02 €) est absorbé sur la TVA collectée ; un écart plus
important (ajustements) génère une ligne d'ajustement explicite et un message.
"""

import json
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# ── Configuration des comptes (modifiable sans toucher au code) ───────────────
# Clé = (plateforme, régime). Le régime « non-france » = ventes hors UE/export :
# pas de TVA. Les numéros de compte France proviennent du modèle Quadra fourni.
#
# ⚠️ Ces valeurs sont des DÉFAUTS. Pour les faire varier sans modifier le code,
# créer un fichier `journal_config.json` à côté de ce module (ou pointer la
# variable d'environnement JOURNAL_CONFIG_FILE vers un JSON). Voir
# `journal_config.example.json`. Le fichier surcharge (deep-merge) ces défauts.
JOURNAL_CONFIG: dict[tuple[str, str], dict] = {
    ("tiktok", "france"): {
        "code_journal": "VTE",
        "libelle": "F TIKTOK - FRANCE",
        "vat_rate": 0.20,
        "comptes": {
            "encaissement": "90TIKTFR",   # Débit  — montant encaissé
            "ventes": "70721000",         # Crédit — ventes HT
            "port": "70800000",           # Crédit — port HT
            "tva_collectee": "44572000",  # Crédit — TVA collectée
            "commission": "62220000",     # Débit  — commission HT
            "tva_deductible": "44566000", # Débit  — TVA déductible sur commission
            # Ajustements : compte d'ATTENTE (classe 47). Un écart non rapproché
            # est parqué ici pour être reclassé MANUELLEMENT par le comptable,
            # jamais fondu dans les ventes/charges (ne pas fausser résultat & TVA).
            "ajustement": "471000",
        },
    },
    # Hors France (export) : pas de TVA. Comptes repris de la France à défaut de
    # modèle dédié — À CONFIRMER (compte de ventes export éventuellement distinct).
    ("tiktok", "non-france"): {
        "code_journal": "VTE",
        "libelle": "F TIKTOK - EXPORT",
        "vat_rate": 0.0,
        "comptes": {
            "encaissement": "90TIKTFR",
            "ventes": "70721000",
            "port": "70800000",
            "tva_collectee": None,
            "commission": "62220000",
            "tva_deductible": None,
            "ajustement": "471000",
        },
    },
}

ROUND_TOLERANCE = 0.02  # écart d'arrondi absorbé silencieusement sur la TVA


def _config_path() -> str | None:
    env = os.environ.get("JOURNAL_CONFIG_FILE")
    if env:
        return env
    default = Path(__file__).with_name("journal_config.json")
    return str(default) if default.exists() else None


def _effective_config() -> dict[tuple[str, str], dict]:
    """Défauts surchargés par le fichier de configuration externe, s'il existe.

    Format JSON attendu : clés « plateforme|régime » (ex. "tiktok|france"),
    valeurs = sous-ensemble de {code_journal, libelle, vat_rate, comptes}.
    Le bloc `comptes` est fusionné compte par compte (on peut n'en changer qu'un).
    """
    base = {k: {**v, "comptes": dict(v["comptes"])} for k, v in JOURNAL_CONFIG.items()}
    path = _config_path()
    if not path:
        return base
    try:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return base  # config illisible → on garde les défauts (jamais d'échec dur)
    if not isinstance(data, dict):
        return base
    for raw_key, override in data.items():
        try:
            platform, regime = (p.strip().lower() for p in str(raw_key).split("|", 1))
        except ValueError:
            continue
        tkey = (platform, regime)
        entry = base.setdefault(tkey, {"code_journal": "VTE", "libelle": "", "vat_rate": 0.0, "comptes": {}})
        if isinstance(override, dict):
            for k, v in override.items():
                if k == "comptes" and isinstance(v, dict):
                    entry["comptes"].update(v)
                else:
                    entry[k] = v
    return base


def _r2(x) -> float:
    return round(float(x), 2)


def country_key(country: str) -> str:
    return "france" if str(country).strip().lower() == "france" else "non-france"


def get_config(platform: str, country: str) -> dict:
    cfg = _effective_config().get((platform, country_key(country)))
    if cfg is None:
        raise ValueError(
            f"Aucune configuration de journal pour la plateforme « {platform} » "
            f"(régime {country_key(country)})."
        )
    return cfg


def _emit(lines: list[dict], *, code: str, date: str, compte: str, libelle: str,
          amount: float, side: str) -> None:
    """Ajoute une ligne d'écriture. Ignore les montants nuls ; un montant négatif
    bascule automatiquement sur la colonne opposée (Débit ↔ Crédit)."""
    amount = _r2(amount)
    if amount == 0:
        return
    if amount < 0:
        side = "credit" if side == "debit" else "debit"
        amount = -amount
    lines.append({
        "code": code,
        "date": date,
        "compte": compte,
        "libelle": libelle,
        "debit": amount if side == "debit" else None,
        "credit": amount if side == "credit" else None,
    })


def _line_totals(lines: list[dict]) -> tuple[float, float]:
    d = _r2(sum(l["debit"] or 0 for l in lines))
    c = _r2(sum(l["credit"] or 0 for l in lines))
    return d, c


def build_journal(df: pd.DataFrame, country: str, platform: str = "tiktok") -> tuple[list[dict], list[str]]:
    """Construit les lignes du journal à partir du DataFrame nettoyé (`clean_data`).

    Renvoie `(lignes, notes)` où `notes` liste les avertissements éventuels
    (ajustements, hypothèses sur le régime export…).
    """
    cfg = get_config(platform, country)
    comptes = cfg["comptes"]
    rate = cfg["vat_rate"]
    libelle = cfg["libelle"]
    code = cfg["code_journal"]
    notes: list[str] = []

    if country_key(country) != "france":
        notes.append(
            "Régime hors France : aucune TVA appliquée. Les comptes export sont "
            "repris du paramétrage France par défaut — à confirmer avec le plan comptable."
        )

    df = df.copy()
    for col in ["net_sales", "shipping", "fees", "adjustments", "total_settlement_amount"]:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    df["_day"] = pd.to_datetime(df["date"], errors="coerce")
    n_bad = int(df["_day"].isna().sum())
    if n_bad:
        notes.append(f"{n_bad} ligne(s) ignorée(s) : date illisible.")
    df = df.dropna(subset=["_day"])

    grouped = (
        df.groupby(df["_day"].dt.normalize(), sort=True)[
            ["net_sales", "shipping", "fees", "adjustments", "total_settlement_amount"]
        ]
        .sum()
        .reset_index()
    )

    lines: list[dict] = []
    for _, row in grouped.iterrows():
        day = row["_day"].strftime("%d/%m/%Y")
        net, ship, fees = row["net_sales"], row["shipping"], row["fees"]
        adj, settlement = row["adjustments"], row["total_settlement_amount"]

        # Décomposition HT / TVA (la TVA est l'écart TTC − HT, pour tomber juste).
        ht_sales = _r2(net / (1 + rate))
        tva_sales = _r2(net - ht_sales)
        ht_port = _r2(ship / (1 + rate))
        tva_port = _r2(ship - ht_port)
        fee_ttc = abs(fees)
        ht_fee = _r2(fee_ttc / (1 + rate))
        tva_fee = _r2(fee_ttc - ht_fee)
        tva_coll = _r2(tva_sales + tva_port)

        day_lines: list[dict] = []
        _emit(day_lines, code=code, date=day, compte=comptes["encaissement"], libelle=libelle, amount=settlement, side="debit")
        _emit(day_lines, code=code, date=day, compte=comptes["ventes"], libelle=libelle, amount=ht_sales, side="credit")
        if comptes["port"]:
            _emit(day_lines, code=code, date=day, compte=comptes["port"], libelle=libelle, amount=ht_port, side="credit")
        if comptes["tva_collectee"]:
            _emit(day_lines, code=code, date=day, compte=comptes["tva_collectee"], libelle=libelle, amount=tva_coll, side="credit")
        if fees != 0:
            _emit(day_lines, code=code, date=day, compte=comptes["commission"], libelle=libelle, amount=ht_fee, side="debit")
            if comptes["tva_deductible"]:
                _emit(day_lines, code=code, date=day, compte=comptes["tva_deductible"], libelle=libelle, amount=tva_fee, side="debit")

        # ── Équilibrage Débit = Crédit ──
        td, tc = _line_totals(day_lines)
        diff = _r2(td - tc)  # >0 : trop de débit → besoin de crédit
        if diff != 0:
            if abs(diff) <= ROUND_TOLERANCE and _r2(adj) == 0:
                _absorb_rounding(day_lines, comptes["tva_collectee"] or comptes["ventes"], diff)
            else:
                _emit(day_lines, code=code, date=day, compte=comptes["ajustement"], libelle=libelle, amount=diff, side="credit")
                notes.append(
                    f"{day} : ligne d'ajustement de {diff:+.2f} € ajoutée sur le compte "
                    f"{comptes['ajustement']} pour équilibrer (ajustements/écart de relevé)."
                )

        lines.extend(day_lines)

    return lines, notes


def _absorb_rounding(day_lines: list[dict], compte: str, diff: float) -> None:
    """Absorbe un écart d'arrondi (≤ 0,02 €) sur une ligne de crédit existante
    (TVA collectée de préférence), pour garantir Débit = Crédit au centime."""
    for l in day_lines:
        if l["compte"] == compte and l["credit"] is not None:
            l["credit"] = _r2(l["credit"] + diff)
            return
    # Pas de ligne adéquate : on en crée une.
    _emit(day_lines, code=day_lines[0]["code"], date=day_lines[0]["date"],
          compte=compte, libelle=day_lines[0]["libelle"], amount=diff, side="credit")


def journal_totals(lines: list[dict]) -> dict:
    d, c = _line_totals(lines)
    return {
        "debit": d,
        "credit": c,
        "balanced": abs(_r2(d - c)) <= 0.01,
        "lines": len(lines),
    }


# ── Export Excel (format épuré, calqué sur le modèle) ─────────────────────────
_HEADERS = ["Code", "Date", "Compte", "Libellé", "Débit", "Crédit"]
_WIDTHS = {"A": 8, "B": 12, "C": 12, "D": 26, "E": 12, "F": 12}


def write_journal_xlsx(lines: list[dict], path, *, entreprise: str = "Journal des ventes",
                       platform: str = "TIKTOK", country: str = "France") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"

    edited = datetime.now().strftime("Édité le %d/%m/%Y à %Hh%M")
    banner = f"Cegid Quadra Comptabilité — {entreprise} ({platform} {country})"
    ws.append([banner, None, None, None, None, edited])
    ws["A1"].font = Font(bold=True)
    ws.append([])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(_HEADERS)
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    for l in lines:
        ws.append([l["code"], l["date"], l["compte"], l["libelle"], l["debit"], l["credit"]])

    for col, w in _WIDTHS.items():
        ws.column_dimensions[col].width = w
    for r in range(header_row + 1, ws.max_row + 1):
        for c in ("E", "F"):
            ws[f"{c}{r}"].number_format = "#,##0.00"

    wb.save(path)
