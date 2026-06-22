"""Tests du générateur de journal d'écritures (format Quadra).

Les montants de référence proviennent du relevé TikTok réel (31/08/2025) :
net_sales=139.90 TTC, shipping=2.92 TTC, fees=-7.00, total_settlement=135.82.
L'invariant central est l'équilibre Débit = Crédit de chaque écriture.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from journal import build_journal, journal_totals, write_journal_xlsx, get_config


def _by_account(lines: list[dict]) -> dict[str, dict]:
    return {l["compte"]: l for l in lines}


@pytest.fixture
def one_day_fr() -> pd.DataFrame:
    return pd.DataFrame({
        "date": ["2025-08-31"],
        "total_settlement_amount": [135.82],
        "net_sales": [139.90],
        "shipping": [2.92],
        "fees": [-7.00],
        "adjustments": [0.0],
    })


def test_france_mapping_and_balance(one_day_fr):
    lines, notes = build_journal(one_day_fr, "France")
    acc = _by_account(lines)

    # Encaissement au débit = total settlement
    assert acc["90TIKTFR"]["debit"] == 135.82
    # Ventes HT au crédit (139.90 / 1.20)
    assert acc["70721000"]["credit"] == 116.58
    # Port HT au crédit (2.92 / 1.20)
    assert acc["70800000"]["credit"] == 2.43
    # TVA collectée = TVA(ventes) + TVA(port)
    assert acc["44572000"]["credit"] == 23.81
    # Commission HT au débit (7.00 / 1.20)
    assert acc["62220000"]["debit"] == 5.83
    # TVA déductible sur commission
    assert acc["44566000"]["debit"] == 1.17

    totals = journal_totals(lines)
    assert totals["balanced"] is True
    assert totals["debit"] == totals["credit"] == 142.82
    assert notes == []


def test_non_france_has_no_vat_and_balances(one_day_fr):
    lines, notes = build_journal(one_day_fr, "Allemagne")
    accounts = {l["compte"] for l in lines}
    # Pas de comptes de TVA hors France
    assert "44572000" not in accounts
    assert "44566000" not in accounts
    acc = _by_account(lines)
    assert acc["90TIKTFR"]["debit"] == 135.82
    assert acc["70721000"]["credit"] == 139.90   # plein TTC = HT (pas de TVA)
    assert acc["62220000"]["debit"] == 7.00
    assert journal_totals(lines)["balanced"] is True
    assert any("hors France" in n for n in notes)


def test_multi_day_aggregation_balances():
    df = pd.DataFrame({
        "date": ["2025-08-01", "2025-08-01", "2025-08-02"],
        "total_settlement_amount": [54.12, 88.37, 190.38],
        "net_sales": [30.00, 50.00, 200.00],
        "shipping": [2.92, 0.00, 5.00],
        "fees": [-3.00, -4.00, -10.00],
        "adjustments": [0.0, 0.0, 0.0],
    })
    lines, _ = build_journal(df, "France")
    # 2 jours distincts → 2 lignes d'encaissement
    enc = [l for l in lines if l["compte"] == "90TIKTFR"]
    assert len(enc) == 2
    assert journal_totals(lines)["balanced"] is True


def test_adjustment_goes_to_waiting_account_with_note():
    df = pd.DataFrame({
        "date": ["2025-08-05"],
        "total_settlement_amount": [110.00],   # 10 € de plus que net+ship+fees
        "net_sales": [100.00],
        "shipping": [0.0],
        "fees": [0.0],
        "adjustments": [10.0],
    })
    lines, notes = build_journal(df, "France")
    assert journal_totals(lines)["balanced"] is True
    assert any("ajustement" in n.lower() for n in notes)
    # L'écart est parqué sur un compte d'attente (classe 47), pas sur les ventes.
    assert "471000" in {l["compte"] for l in lines}


def test_negative_settlement_flips_side():
    # Jour de remboursement net : encaissement négatif → bascule au crédit.
    df = pd.DataFrame({
        "date": ["2025-08-06"],
        "total_settlement_amount": [-50.00],
        "net_sales": [-60.00],
        "shipping": [0.0],
        "fees": [10.0],
        "adjustments": [0.0],
    })
    lines, _ = build_journal(df, "France")
    enc = _by_account(lines)["90TIKTFR"]
    assert enc["debit"] is None and enc["credit"] == 50.00
    assert journal_totals(lines)["balanced"] is True


def test_write_xlsx_structure(one_day_fr, tmp_path: Path):
    lines, _ = build_journal(one_day_fr, "France")
    out = tmp_path / "journal.xlsx"
    write_journal_xlsx(lines, out, platform="TIKTOK", country="France")
    assert out.is_file()

    wb = load_workbook(out)
    ws = wb.active
    assert ws["A1"].value.startswith("Cegid Quadra Comptabilité")
    # Ligne d'en-tête des colonnes
    header = [c.value for c in ws[4]]
    assert header[:6] == ["Code", "Date", "Compte", "Libellé", "Débit", "Crédit"]
    # Première ligne de données : journal VTE, date formatée
    assert ws["A5"].value == "VTE"
    assert ws["B5"].value == "31/08/2025"


def test_unknown_platform_raises(one_day_fr):
    with pytest.raises(ValueError):
        build_journal(one_day_fr, "France", platform="amazon")


def test_get_config_france():
    cfg = get_config("tiktok", "France")
    assert cfg["vat_rate"] == 0.20
    assert cfg["comptes"]["encaissement"] == "90TIKTFR"


def test_external_config_override(one_day_fr, tmp_path, monkeypatch):
    """Un fichier JSON externe surcharge les comptes sans toucher au code."""
    import json as _json
    cfg_file = tmp_path / "journal_config.json"
    cfg_file.write_text(_json.dumps({
        "tiktok|france": {"comptes": {"encaissement": "90CUSTOM", "ventes": "70709999"}}
    }), encoding="utf-8")
    monkeypatch.setenv("JOURNAL_CONFIG_FILE", str(cfg_file))

    cfg = get_config("tiktok", "France")
    assert cfg["comptes"]["encaissement"] == "90CUSTOM"   # surchargé
    assert cfg["comptes"]["ventes"] == "70709999"          # surchargé
    assert cfg["comptes"]["tva_collectee"] == "44572000"   # défaut conservé

    lines, _ = build_journal(one_day_fr, "France")
    accounts = {l["compte"] for l in lines}
    assert "90CUSTOM" in accounts and "70709999" in accounts
    assert journal_totals(lines)["balanced"] is True
